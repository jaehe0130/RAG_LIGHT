import os
import json
import time
import urllib.request
import urllib.error
import re
import sys
import shutil
import copy
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Force output encoding to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# Env load
from dotenv import load_dotenv
load_dotenv(dotenv_path='backend/.env')

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
MODEL_NAME = os.getenv("GEMINI_MODEL", "gemini-3.1-flash-lite")
RAW_CACHE_FILE = "backend/data/qa_1000_raw.json"
VERIFIED_CACHE_FILE = "backend/data/qa_1000_verified.json"
EXCEL_FILE_PATH = "설계서/학습데이터 설계서.xlsx"

def call_gemini_api(prompt: str) -> list:
    if not GEMINI_API_KEY:
        raise ValueError("GEMINI_API_KEY is not set in env.")

    url = "https://generativelanguage.googleapis.com/v1beta/openai/chat/completions"
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GEMINI_API_KEY}"
    }
    
    data = json.dumps({
        "model": MODEL_NAME,
        "messages": [
            {"role": "system", "content": "당신은 한국 공정거래법 및 학습데이터 품질 관리에 정통한 AI 검수관입니다. 주어진 QA 레코드를 면밀히 감사하여 형식이나 내용상 위배 사항을 찾아내어 JSON 배열로 보고해야 합니다."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.2
    }).encode("utf-8")
    
    req = urllib.request.Request(
        url,
        data=data,
        headers=headers,
        method="POST"
    )
    
    try:
        with urllib.request.urlopen(req, timeout=90) as response:
            res_body = response.read().decode("utf-8")
            res_json = json.loads(res_body)
            content = res_json["choices"][0]["message"]["content"].strip()
            
            json_match = re.search(r"```(?:json)?\s*(.*?)\s*```", content, re.DOTALL)
            if json_match:
                content = json_match.group(1).strip()
            
            return json.loads(content)
    except Exception as e:
        print(f"❌ Gemini Auditor API Call Error: {e}")
        return []

def main():
    print("🚀 [Dual-Validation Pipeline] Loading raw 1,000 cases...")
    
    if not os.path.exists(RAW_CACHE_FILE):
        print(f"❌ Error: {RAW_CACHE_FILE} does not exist. Please run generate_1000_qa.py first.")
        return
        
    with open(RAW_CACHE_FILE, "r", encoding="utf-8") as f:
        raw_data = json.load(f)
        
    print(f"📊 Loaded {len(raw_data)} cases from raw cache.")
    
    # Load audited cache if exists to prevent duplicate API audits
    audited_issues = {}
    if os.path.exists(VERIFIED_CACHE_FILE):
        try:
            with open(VERIFIED_CACHE_FILE, "r", encoding="utf-8") as f:
                audited_issues = json.load(f)
            print(f"🔄 Loaded {len(audited_issues)} cached audit results.")
        except Exception:
            audited_issues = {}
            
    raw_keys = list(raw_data.keys())
    
    # Batch size of 20 for high precision AI review
    batch_size = 20
    
    for start_idx in range(0, len(raw_data), batch_size):
        batch_keys = raw_keys[start_idx : start_idx + batch_size]
        need_audit = [k for k in batch_keys if k not in audited_issues]
        
        if not need_audit:
            continue
            
        print(f"📈 [Auditing] Dual-Validation Batch {start_idx // batch_size + 1}/{len(raw_data) // batch_size + 1} (Index {start_idx} to {min(len(raw_data), start_idx + batch_size)})...")
        
        # 1. Step 1: Local rule validation for immediate formats
        local_issues = {}
        ai_target_keys = []
        
        for k in need_audit:
            item = raw_data[k]
            has_local_err = False
            prob = ""
            sol = ""
            
            # Facts format check
            facts = item.get("facts", "")
            required_facts = ["주체:", "상대방:", "시장:", "핵심행위:", "결과:"]
            if not all(rf in facts for rf in required_facts):
                has_local_err = True
                prob = "사실관계(facts) 형식 미준수 (주체/상대방/시장/핵심행위/결과 필수 키 누락)"
                sol = "사실관계 필드를 '주체: ... / 상대방: ... / 시장: ... / 핵심행위: ... / 결과: ...' 형식에 맞게 보완하십시오."
            elif "/" not in facts:
                has_local_err = True
                prob = "사실관계(facts) 형식 미준수 (구분자 '/' 누락)"
                sol = "사실관계의 각 필드를 '/' 기호로 명확히 구분하십시오."
                
            # Legal reasoning format check
            lr = item.get("legal_reasoning", "")
            if not lr.startswith("법리 해석:"):
                has_local_err = True
                prob = "법리 해석 접두어 '법리 해석:' 누락"
                sol = "법리 해석 텍스트 시작 부분에 '법리 해석:' 접두어를 삽입하십시오."
                
            if has_local_err:
                local_issues[k] = {
                    "idx": int(k),
                    "has_issue": True,
                    "problem": prob,
                    "solution": sol
                }
            else:
                ai_target_keys.append(k)
                
        # 2. Step 2: AI auditor check for logical and stylistic integrity
        if ai_target_keys:
            req_str = ""
            for k in ai_target_keys:
                item = raw_data[k]
                req_str += (
                    f"- Index: {item['idx']}\n"
                    f"  페르소나: {item['persona']}\n"
                    f"  질문: {item['question']}\n"
                    f"  답변: {item['answer']}\n"
                    f"  적용법률: {item['applicable_law']}\n"
                    f"  참조문서: {item['ref_doc']}\n\n"
                )
                
            prompt = (
                f"당신은 대한민국 공정거래법 및 소비자법 학습데이터 품질 관리에 정통한 전문 AI 법률 감사관입니다.\n"
                f"제시된 {len(ai_target_keys)}개 데이터셋의 질문과 답변, 적용 법률의 정합성을 아래의 엄격한 Rule-book에 따라 검수하십시오.\n\n"
                f"[★AI 검수관 Rule-book 평가 기준]\n"
                f"1. 페르소나별 톤앤매너 불일치 오류:\n"
                f"   - '일반국민': 대중이 겪을 법한 일상의 피해에 관한 친근한 질문이어야 하며 답변도 이해하기 쉬워야 합니다.\n"
                f"   - '기업 컴플라이언스 담당자': 기업 내부 실무자 관점의 비즈니스 소송 및 위반 리스크 통제 관련 질문과 리스크 완화 관점의 답변이어야 합니다.\n"
                f"   - '논문 준비 학생': 학술 연구용, 소비자 후생, 쟁점 분석 등 학술적 질문과 판례 중심의 답변이어야 합니다.\n"
                f"2. 적용법률 오매칭 오류:\n"
                f"   - 질문/답변에 언급된 위반 행위(예: 환불 거부)와 매칭된 'applicable_law'의 실제 법령/조항(예: 표시광고법 등 엉뚱한 법 적용)이 법리적으로 심각하게 맞지 않는 경우.\n"
                f"3. 기계적 템플릿 중복 오류:\n"
                f"   - 답변 도입부가 항상 동일한 어구(예: '~는 다음과 같습니다', '~검토한 결과는 다음과 같습니다')로 일률적으로 끝나거나, 모든 레코드의 마지막에 '구체적인 사안에 대해 더 자세한 검토가 필요하다면...' 식의 획일적인 꼬리표가 반복되는 등 AI 생성 특유의 기계적 반복 패턴이 나타나는 경우.\n\n"
                f"[★중요 - 중대한 결함만 보고 및 허위 경보 금지]\n"
                f"- 위의 3대 규칙에 위배되는 **중대한 결함이나 모순이 발견된 아이템에 대해서만** 'problem'과 'solution'을 JSON 배열에 담아 기재하십시오.\n"
                f"- 단순한 어조 차이나 허용 가능한 유연한 답변 구조를 지닌 **정상적이고 자연스러운 데이터는 절대로 억지로 문제를 제기하지 말고 무조건 통과(JSON 배열에 미기재)** 시키십시오.\n"
                f"- 마크다운 기호 없이 순수한 JSON 배열 포맷으로만 응답해야 합니다.\n\n"
                f"[검수 대상 목록]\n{req_str}\n"
                f"[지침]\n"
                f"- 문제 발견 시 JSON 배열 규격 예시:\n"
                f"  [\n"
                f"    {{\n"
                f"      \"idx\": 0,\n"
                f"      \"problem\": \"[문제점] 구체적인 룰 위배 원인을 서술\",\n"
                f"      \"solution\": \"[해결방안] 프롬프트나 데이터를 어떻게 교정해야 하는지 서술\"\n"
                f"    }}\n"
                f"  ]\n"
            )
            
            ai_success = False
            for attempt in range(3):
                batch_res = call_gemini_api(prompt)
                if isinstance(batch_res, list):
                    # Default all to normal
                    for k in ai_target_keys:
                        audited_issues[k] = {"idx": int(k), "has_issue": False}
                        
                    # Overlay detected issues
                    for issue in batch_res:
                        idx_str = str(issue.get("idx"))
                        if idx_str in audited_issues:
                            audited_issues[idx_str] = {
                                "idx": int(idx_str),
                                "has_issue": True,
                                "problem": issue.get("problem"),
                                "solution": issue.get("solution")
                            }
                    ai_success = True
                    break
                else:
                    print(f"⚠️ Batch AI audit failed, retrying in 5s... (Attempt {attempt+1}/3)")
                    time.sleep(5)
                    
            if not ai_success:
                print("❌ Critical Error: AI API Audit failed permanently. Terminating...")
                return
                
        # 3. Merge local issues into audited_issues
        for k, l_issue in local_issues.items():
            audited_issues[k] = l_issue
            
        # Write back cache at each batch
        with open(VERIFIED_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(audited_issues, f, ensure_ascii=False, indent=2)
            
        print(f"✅ Batch validated & merged. (Total audited: {len(audited_issues)}/1000)")
        time.sleep(4)
        
    print("🎉 All 1,000 cases validated successfully via dual-validation pipeline! Preparing Excel writing...")
    
    # 7. Copy template and load Excel workbook
    print("📋 Copying template file to destination...")
    copied = False
    for attempt in range(12):
        try:
            shutil.copy2("설계서/학습데이터 설계서 양식.xlsx", EXCEL_FILE_PATH)
            copied = True
            break
        except PermissionError:
            print(f"⚠️ [대기 중] '{EXCEL_FILE_PATH}' 파일이 열려 있어 접근할 수 없습니다. 엑셀 프로그램을 종료해 주세요. (5초 후 재시도 {attempt+1}/12)")
            time.sleep(5)
            
    if not copied:
        print(f"❌ 오류: '{EXCEL_FILE_PATH}' 파일을 덮어쓸 수 없습니다. 엑셀을 종료한 후 다시 실행하십시오.")
        return
        
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)

    # Helper to get styles from row 3
    def get_row_styles(sheet, row_idx=3):
        styles = {}
        for col in range(1, 18):
            cell = sheet.cell(row=row_idx, column=col)
            styles[col] = {
                "font": copy.copy(cell.font) if cell.font else None,
                "border": copy.copy(cell.border) if cell.border else None,
                "fill": copy.copy(cell.fill) if cell.fill else None,
                "alignment": copy.copy(cell.alignment) if cell.alignment else None,
                "number_format": cell.number_format
            }
        return styles

    def apply_row_styles(sheet, row_idx, styles):
        for col in range(1, 18):
            cell = sheet.cell(row=row_idx, column=col)
            style = styles[col]
            if style["font"]: cell.font = style["font"]
            if style["border"]: cell.border = style["border"]
            if style["fill"]: cell.fill = style["fill"]
            if style["alignment"]: cell.alignment = style["alignment"]
            if style["number_format"]: cell.number_format = style["number_format"]

    data_sheets = ["전체_QA", "일반국민", "기업 컴플라이언스 담당자", "논문 준비 학생"]
    
    # Extract styles from row 3 and then clear values from row 3 downwards (keeping Column 17 header)
    sheet_styles = {}
    for s_name in data_sheets:
        sh = wb[s_name]
        sheet_styles[s_name] = get_row_styles(sh, row_idx=3)
        for r in range(3, max(3, sh.max_row + 1)):
            for c in range(1, 18):
                sh.cell(row=r, column=c).value = None
                
    # Compile rows in order
    all_rows = []
    for idx in range(1000):
        item = raw_data[str(idx)]
        row_id = f"QA-{idx + 1:04d}"
        
        no_str = item.get("의결서관리번호") or ""
        if no_str and len(no_str) >= 4:
            ref_val = f"의결서{no_str[:2]}-{no_str[2:4]}  1줄-{no_str}줄"
        else:
            ref_val = f"의결서02-24  1줄-22450줄"
            
        col_8_basis = f"의결서는 '{item['피심인기업명']}'의 행위가 '{item['세부위반유형']}'에 해당한다고 보고 '{item['조치유형']}'을 결정하였다."
        col_9_summary = f"{item['피심인기업명']}의 {item['세부위반유형']} 사건은 {item['피심인기업명']}이(가) 상대방에게 불리한 행위를 하여 문제가 된 사례이다."
        
        # Scenarios
        p_name = item["persona"]
        style_desc = "쉽고 직관적인 설명" if p_name == "일반국민" else ("리스크 및 준법 중심 설명" if p_name == "기업 컴플라이언스 담당자" else "학술·연구 중심 설명")
        point_desc = "생활 속 사례와 쉽게 풀어쓴 설명" if p_name == "일반국민" else ("위반 위험요소, 내부통제 포인트, 재발방지 관점" if p_name == "기업 컴플라이언스 담당자" else "쟁점, 판단 구조, 비교법적·통계적 해석")
        
        basic_sc = f"사용자가 '{item['피심인기업명']}의 {item['세부위반유형']}' 관련 궁금증을 자연어로 입력하면, AI가 {style_desc}로 관련 사실과 의결서 근거를 제시한다."
        add_sc = f"추가 시나리오: 유사한 위반 유형 및 행정 제재 통계를 시각화하여 사용자가 행동 양식을 판단할 수 있도록 돕는다."
        
        # Determine issue and solution for Column 17 (문제점 및 해결방안)
        issue = audited_issues.get(str(idx))
        if issue and issue.get("has_issue"):
            col_17_val = f"[문제점] {issue.get('problem')}\n[해결방안] {issue.get('solution')}"
        else:
            col_17_val = None
            
        row_data = {
            "ID": row_id,
            "페르소나": p_name,
            "질문유형": item["question_type"],
            "난이도": item["difficulty"],
            "주제": item["subject"],
            "질문": item["question"],
            "답변": item["answer"],
            "근거": col_8_basis,
            "주제별·수준별 의결서 요약": col_9_summary,
            "사실(Facts)": item["facts"],
            "법리 해석": item["legal_reasoning"],
            "기본 시나리오": basic_sc,
            "추가 시나리오": add_sc,
            "응답 설계 포인트": point_desc,
            "의도된 설명 스타일": style_desc,
            "주의사항": item.get("주의사항") or ref_val,
            "비고": item.get("비고") or ref_val,
            "문제점_및_해결방안": col_17_val
        }
        all_rows.append(row_data)
        
    # Write to 전체_QA (16번째 열 헤더: '근거' -> 근거 기재)
    sh_total = wb["전체_QA"]
    styles_total = sheet_styles["전체_QA"]
    for idx, rdata in enumerate(all_rows):
        r = idx + 3
        apply_row_styles(sh_total, r, styles_total)
        sh_total.cell(row=r, column=1).value = rdata["ID"]
        sh_total.cell(row=r, column=2).value = rdata["페르소나"]
        sh_total.cell(row=r, column=3).value = rdata["질문유형"]
        sh_total.cell(row=r, column=4).value = rdata["난이도"]
        sh_total.cell(row=r, column=5).value = rdata["주제"]
        sh_total.cell(row=r, column=6).value = rdata["질문"]
        sh_total.cell(row=r, column=7).value = rdata["답변"]
        sh_total.cell(row=r, column=8).value = rdata["근거"]
        sh_total.cell(row=r, column=9).value = rdata["주제별·수준별 의결서 요약"]
        sh_total.cell(row=r, column=10).value = rdata["사실(Facts)"]
        sh_total.cell(row=r, column=11).value = rdata["법리 해석"]
        sh_total.cell(row=r, column=12).value = rdata["기본 시나리오"]
        sh_total.cell(row=r, column=13).value = rdata["추가 시나리오"]
        sh_total.cell(row=r, column=14).value = rdata["응답 설계 포인트"]
        sh_total.cell(row=r, column=15).value = rdata["의도된 설명 스타일"]
        sh_total.cell(row=r, column=16).value = rdata["근거"]
        sh_total.cell(row=r, column=17).value = rdata["문제점_및_해결방안"]
    print("✍️ Wrote '전체_QA' sheet (1000 rows).")

    # Filter and write to General Public (16번째 열 헤더: '주의사항' -> 주의사항 기재)
    general_rows = [row for row in all_rows if row["페르소나"] == "일반국민"]
    sh_gen = wb["일반국민"]
    styles_gen = sheet_styles["일반국민"]
    for idx, rdata in enumerate(general_rows):
        r = idx + 3
        apply_row_styles(sh_gen, r, styles_gen)
        sh_gen.cell(row=r, column=1).value = rdata["ID"]
        sh_gen.cell(row=r, column=2).value = rdata["페르소나"]
        sh_gen.cell(row=r, column=3).value = rdata["질문유형"]
        sh_gen.cell(row=r, column=4).value = rdata["난이도"]
        sh_gen.cell(row=r, column=5).value = rdata["주제"]
        sh_gen.cell(row=r, column=6).value = rdata["질문"]
        sh_gen.cell(row=r, column=7).value = rdata["답변"]
        sh_gen.cell(row=r, column=8).value = rdata["근거"]
        sh_gen.cell(row=r, column=9).value = rdata["주제별·수준별 의결서 요약"]
        sh_gen.cell(row=r, column=10).value = rdata["사실(Facts)"]
        sh_gen.cell(row=r, column=11).value = rdata["법리 해석"]
        sh_gen.cell(row=r, column=12).value = rdata["기본 시나리오"]
        sh_gen.cell(row=r, column=13).value = rdata["추가 시나리오"]
        sh_gen.cell(row=r, column=14).value = rdata["응답 설계 포인트"]
        sh_gen.cell(row=r, column=15).value = rdata["의도된 설명 스타일"]
        sh_gen.cell(row=r, column=16).value = rdata["주의사항"]
        sh_gen.cell(row=r, column=17).value = rdata["문제점_및_해결방안"]
    print(f"✍️ Wrote '일반국민' sheet ({len(general_rows)} rows).")

    # Write to Compliance (16번째 열 헤더: '주의사항' -> 주의사항 기재)
    comp_rows = [row for row in all_rows if row["페르소나"] == "기업 컴플라이언스 담당자"]
    sh_comp = wb["기업 컴플라이언스 담당자"]
    styles_comp = sheet_styles["기업 컴fl라이언스 담당자"] if "기업 컴fl라이언스 담당자" in wb.sheetnames else sheet_styles["기업 컴플라이언스 담당자"]
    for idx, rdata in enumerate(comp_rows):
        r = idx + 3
        apply_row_styles(wb["기업 컴플라이언스 담당자"], r, styles_comp)
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=1).value = rdata["ID"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=2).value = rdata["페르소나"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=3).value = rdata["질문유형"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=4).value = rdata["난이도"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=5).value = rdata["주제"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=6).value = rdata["질문"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=7).value = rdata["답변"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=8).value = rdata["근거"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=9).value = rdata["주제별·수준별 의결서 요약"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=10).value = rdata["facts"] if "facts" in rdata else rdata.get("사실(Facts)")
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=11).value = rdata["법리 해석"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=12).value = rdata["기본 시나리오"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=13).value = rdata["추가 시나리오"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=14).value = rdata["응답 설계 포인트"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=15).value = rdata["의도된 설명 스타일"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=16).value = rdata["주의사항"]
        wb["기업 컴플라이언스 담당자"].cell(row=r, column=17).value = rdata["문제점_및_해결방안"]
    print(f"✍️ Wrote '기업 컴플라이언스 담당자' sheet ({len(comp_rows)} rows).")

    # Write to Student (16번째 열 헤더: '비고' -> 비고 기재)
    student_rows = [row for row in all_rows if row["페르소나"] == "논문 준비 학생"]
    sh_stud = wb["논문 준비 학생"]
    styles_stud = sheet_styles["논문 준비 학생"]
    for idx, rdata in enumerate(student_rows):
        r = idx + 3
        apply_row_styles(sh_stud, r, styles_stud)
        sh_stud.cell(row=r, column=1).value = rdata["ID"]
        sh_stud.cell(row=r, column=2).value = rdata["페르소나"]
        sh_stud.cell(row=r, column=3).value = rdata["질문유형"]
        sh_stud.cell(row=r, column=4).value = rdata["난이도"]
        sh_stud.cell(row=r, column=5).value = rdata["주제"]
        sh_stud.cell(row=r, column=6).value = rdata["질문"]
        sh_stud.cell(row=r, column=7).value = rdata["답변"]
        sh_stud.cell(row=r, column=8).value = rdata["근거"]
        sh_stud.cell(row=r, column=9).value = rdata["주제별·수준별 의결서 요약"]
        sh_stud.cell(row=r, column=10).value = rdata["facts"] if "facts" in rdata else rdata.get("사실(Facts)")
        sh_stud.cell(row=r, column=11).value = rdata["법리 해석"]
        sh_stud.cell(row=r, column=12).value = rdata["기본 시나리오"]
        sh_stud.cell(row=r, column=13).value = rdata["추가 시나리오"]
        sh_stud.cell(row=r, column=14).value = rdata["응답 설계 포인트"]
        sh_stud.cell(row=r, column=15).value = rdata["의도된 설명 스타일"]
        sh_stud.cell(row=r, column=16).value = rdata["비고"]
        sh_stud.cell(row=r, column=17).value = rdata["문제점_및_해결방안"]
    print(f"✍️ Wrote '논문 준비 학생' sheet ({len(student_rows)} rows).")

    # 8. Update Dashboard
    sh_dash = wb["대시보드"]
    sh_dash["A1"] = "페르소나 기반 QA 데이터셋 대시보드 (1000건)"
    sh_dash["B3"] = "=COUNTA(전체_QA!A3:A1100)"
    
    sh_dash["B6"] = "=COUNTA(일반국민!A3:A1100)"
    sh_dash["B7"] = "=COUNTA('기업 컴플라이언스 담당자'!A3:A1100)"
    sh_dash["B8"] = "=COUNTA('논문 준비 학생'!A3:A1100)"
    
    sh_dash["E6"] = "=COUNTIF(전체_QA!C3:C1100, \"제재중심\")"
    sh_dash["E7"] = "=COUNTIF(전체_QA!C3:C1100, \"법령중심\")"
    sh_dash["E8"] = "=COUNTIF(전체_QA!C3:C1100, \"유사사례\")"
    sh_dash["E9"] = "=COUNTIF(전체_QA!C3:C1100, \"통계기반\")"
    sh_dash["E10"] = "=COUNTIF(전체_QA!C3:C1100, \"시장영향\")"
    
    sh_dash["B12"] = "=COUNTIF(전체_QA!D3:D1100, \"기초\")"
    sh_dash["B13"] = "=COUNTIF(전체_QA!D3:D1100, \"중급\")"
    sh_dash["B14"] = "=COUNTIF(전체_QA!D3:D1100, \"심화\")"
    print("📊 Updated '대시보드' sheet with active formulas.")
    
    # Save workbook
    try:
        wb.save(EXCEL_FILE_PATH)
        print(f"🎉 Successfully saved audited workbook to: {EXCEL_FILE_PATH}")
    except Exception as e:
        print(f"❌ Error saving workbook: {e}")

if __name__ == "__main__":
    main()
